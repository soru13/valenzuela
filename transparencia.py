import os 
from xlrd import open_workbook
from openpyxl import Workbook, load_workbook

path = "/app/excels"

os.chdir(path)
# wb = Workbook()
#Abrimos el archivo para leerlo
transparencia_excel = load_workbook("{}/{}".format(path,'transparencia.xlsx'))
sheet_transparencia_excel = transparencia_excel["Sheet"]
sheet_transparencia_excel_2 = transparencia_excel["Sheet2"]

# sheets = transparencia_excel.sheetnames
# grab the active worksheet

# ws = wb.active
# wb.save("transparencia.xlsx")
# wb.create_sheet('Sheet2')
# wb.save('transparencia.xlsx')
def fila_iniciar(sheet, row_count):
    for row in range(1, row_count):
        cell = sheet.cell(row, 0)
        if ("ID" == cell.value):
            return row
def isRelationship(sheet, salto_row, que_columna):
    
    for cur_row in range(salto_row+1, salto_row+2):
        cell = sheet.cell(cur_row, que_columna)
        if isinstance(cell.value, str):
            if (cell.value.strip().replace('.','',1).isnumeric()):
                return True
            else:
                return False
        else:
            if (isinstance(cell.value,float) or isinstance(cell.value,int)):
                return True
            else:
                return False
                
def insertar(row_count, cur_col,maxima_row,sheet, salto_row, que_columna, sheet2, is_number, file, name_institute):
    for cur_row in range(salto_row+1, row_count):
        cell = sheet.cell(cur_row, cur_col)
        if cell.value:
            if(sheet2):
                transparencia_excel.active = 1
                if (is_number):
                    if (isinstance(cell.value,float) or isinstance(cell.value,int)):
                        sheet_transparencia_excel_2.cell(row=(maxima_row+cur_row-salto_row), column=que_columna, value=cell.value)
                    else:
                        if (cell.value.strip().replace('.','',1).isnumeric()):
                            sheet_transparencia_excel_2.cell(row=(maxima_row+cur_row-salto_row), column=que_columna, value=float(cell.value))
                        else:
                            sheet_transparencia_excel_2.cell(row=(maxima_row+cur_row-salto_row), column=que_columna, value=cell.value.strip())
                else:
                    sheet_transparencia_excel_2.cell(row=(maxima_row+cur_row-salto_row), column=que_columna, value=cell.value.strip())                        
            else:
                transparencia_excel.active = 0
                if (isinstance(cell.value,float) or isinstance(cell.value,int)):
                    sheet_transparencia_excel.cell(row=(maxima_row+cur_row-salto_row), column=que_columna, value=cell.value)
                else:
                    if (cell.value.strip().replace('.','',1).isnumeric()):
                        sheet_transparencia_excel.cell(row=(maxima_row+cur_row-salto_row), column=que_columna, value=float(cell.value))
                    else:
                        sheet_transparencia_excel.cell(row=(maxima_row+cur_row-salto_row), column=que_columna, value=cell.value.strip())
                sheet_transparencia_excel.cell(row=(maxima_row+cur_row-salto_row), column=46, value=file)
                sheet_transparencia_excel.cell(row=(maxima_row+cur_row-salto_row), column=43, value=name_institute)

def iter_another_table(book,name_table):
    sh = book.sheet_by_name(name_table)
    row_count = sh.nrows
    print('hoja secundaria')
    print(row_count)
    col_count = sh.ncols
    maxima_row = sheet_transparencia_excel_2.max_row
    row = fila_iniciar(sh, row_count)
    iter_columnas(col_count , sh, maxima_row, row_count, row, book, True, '', '')
def relation(cell, book, texto_tabla, is_tabla):
    transparencia_excel.active = 1
    if (is_tabla):
        iter_another_table(book, texto_tabla)
    else:
        tabla = cell.value.split(texto_tabla, 1)[1].replace(")","").strip().upper().replace("_","")
        iter_another_table(book, tabla)

def iter_columnas(col_count , sheet, maxima_row, row_count, row, book, sheet2, file, name_institute):
    for cur_col in range(0, col_count):
        transparencia_excel.active = 0
        cell = sheet.cell(row, cur_col)
        if ('ID' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 1, sheet2, True, file, name_institute)
        if (' Ejercicio' == cell.value or 'EJERCICIO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 2, sheet2, True, file, name_institute)
        if (' Tipo de Procedimiento (catálogo)' == cell.value or ' Tipo de Procedimiento' == cell.value or 'TIPO DE PROCEDIMIENTO (CATÁLOGO)' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 3, sheet2, False, file, name_institute)
        if (' Materia' == cell.value or ' Materia O Tipo de Contratación (catálogo)' == cell.value or ' Categoría:' == cell.value or 'MATERIA (CATÁLOGO)' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 4, sheet2, False, file, name_institute)
        if (' Número de Expediente, Folio O Nomenclatura' == cell.value or 'NÚMERO DE EXPEDIENTE  FOLIO O NOMENCLATURA QUE LO IDENTIFIQUE' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 5, sheet2, False, file, name_institute)
        if (' Nombre(s) Del Adjudicado' == cell.value or ' Nombre(s) Del Contratista O Proveedor' ==  cell.value or ' Nombre(s)' == cell.value or 'NOMBRE(S) DEL ADJUDICADO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 6, sheet2, False, file, name_institute)
        if (' Primer Apellido Del Adjudicado' == cell.value or ' Primer Apellido' == cell.value or ' Primer Apellido Del Contratista O Proveedor' == cell.value or 'PRIMER APELLIDO DEL ADJUDICADO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 7, sheet2, False, file, name_institute)
        if (' Segundo Apellido Del Adjudicado' == cell.value or ' Segundo Apellido' == cell.value or ' Segundo Apellido Del Contratista O Proveedor' == cell.value or 'SEGUNDO APELLIDO DEL ADJUDICADO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 8, sheet2, False, file, name_institute)
        if (' Nombre O Razón Social Del Adjudicado (Tabla' in cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 9, sheet2, True, file, name_institute)
            relation(cell, book, 'Adjudicado (', False)
        if (' Nombre Completo Del O Los Contratista(s) Elegidos (Tabla' in cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 6, sheet2, True, file, name_institute)
            relation(cell, book, 'Elegidos (', False)
        if (' Origen de Los Recursos Públicos (Tabla' in cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 19, sheet2, True, file, name_institute)
            relation(cell, book, 'Públicos (', False)
        if (' Denominación O Razón Social' == cell.value or
            ' Denominación O Razón Social Del Contratista' == cell.value or
            ' Razón Social' == cell.value or
            ' Razón Social Del Contratista O Proveedor' == cell.value or
            ' Nombre O Razón Social Del Adjudicado' == cell.value or
            'RAZÓN SOCIAL DEL ADJUDICADO' == cell.value):
            if(sheet2):
                insertar(row_count, cur_col, maxima_row, sheet, row, 9, sheet2, False, file, name_institute)
            else:
                is_relationship = isRelationship(sheet, row, cur_col)
                if (is_relationship):
                    insertar(row_count, cur_col, maxima_row, sheet, row, 9, sheet2, True, file, name_institute)
                    relation(cell, book, 'TABLA217180', True)
                else:
                    insertar(row_count, cur_col, maxima_row, sheet, row, 9, sheet2, False, file, name_institute)
        if (' Rfc de La Persona Física O Moral Contratista O Proveedor' == cell.value or ' Registro Federal de Contribuyentes (rfc) de La Persona Física O Moral Adjudicada' == cell.value or 'REGISTRO FEDERAL DE CONTRIBUYENTES (RFC) DE LA PERSONA FÍSICA O MORAL ADJUDICADA' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 10, sheet2, False, file, name_institute)
        if (' Área(s) Solicitante' == cell.value or 'ÁREA(S) SOLICITANTE(S)' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 11, sheet2, False, file, name_institute)
        if (' Área(s) Contratante(s)' == cell.value or ' Unidad Administrativa Contratante' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 12, sheet2, False, file, name_institute)
        if (' Número Que Identifique Al Contrato' == cell.value or 'NÚMERO QUE IDENTIFIQUE AL CONTRATO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 13, sheet2, True, file, name_institute)
        if (' Monto Del Contrato sin Impuestos (en Pesos Mex.)' == cell.value or ' Monto Del Contrato sin Impuestos (en Mxn)' == cell.value or
            'MONTO DEL CONTRATO SIN IMPUESTOS INCLUIDOS' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 14, sheet2, True, file, name_institute)
        if (' Monto Total Del Contrato con Impuestos Incluidos' == cell.value or ' Monto Total Del Contrato con Impuestos Incluidos (mxn)' == cell.value or
            'MONTO TOTAL DEL CONTRATO CON IMPUESTOS INCLUIDOS (EXPRESADO EN PESOS MEXICANOS)' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 15, sheet2, True, file, name_institute)
        if (' Tipo de Moneda' == cell.value or 'TIPO DE MONEDA' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 16, sheet2, False, file, name_institute)
        if (' Tipo de Cambio de Referencia, en su Caso' == cell.value or 'TIPO DE CAMBIO DE REFERENCIA  EN SU CASO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 17, sheet2, False, file, name_institute)
        if (' Forma de Pago' == cell.value or 'FORMA DE PAGO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 18, sheet2, False, file, name_institute)
        if (' Origen de Los Recursos Públicos (catálogo)' == cell.value or ' Origen de Los Recursos Públicos' == cell.value or 'ORIGEN DE LOS RECURSOS PÚBLICOS' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 19, sheet2, False, file, name_institute)
        if (' Fuente de Financiamiento' == cell.value or 'FUENTES DE FINANCIAMIENTO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 20, sheet2, False, file, name_institute)
        if (' Lugar Donde Se Realizará La Obra Pública, en su Caso' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 21, sheet2, False, file, name_institute)
        if (' Descripción de Las Obras, Bienes O Servicios' == cell.value or 'DESCRIPCIÓN DE OBRAS  BIENES O SERVICIOS' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 22, sheet2, False, file, name_institute)
        if (' Fecha de La Convocatoria O Invitación' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 23, sheet2, False, file, name_institute)
        if (' Fecha Del Contrato' == cell.value or 'FECHA DEL CONTRATO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 24, sheet2, False, file, name_institute)
        if (' Hipervínculo Al Comunicado de Suspensión, en su Caso' == cell.value or 'HIPERVÍNCULO AL COMUNICADO DE SUSPENSIÓN  RESCISIÓN O TERMINACIÓN ANTICIPADA DEL CONTRATO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 25, sheet2, False, file, name_institute)
        if (' Hipervínculo Al (los) Dictámenes, en su Caso' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 26, sheet2, False, file, name_institute)
        if (' Hipervínculo a La Convocatoria O Invitaciones' == cell.value or ' Hipervínculo a La Convocatoria O Invitaciones Emitidas' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 27, sheet2, False, file, name_institute)
        if (' Hipervínculo Al Fallo de La Junta de Aclaraciones O Al Documento Correspondiente' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 28, sheet2, False, file, name_institute)
        if (' Hipervínculo Al Documento Donde Conste La Presentación Las Propuestas' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 29, sheet2, False, file, name_institute)
        # #obras públicas campo aun faltantes que no tenia
        if (' Descripción de Las Razones Que Justifican su Elección' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 30, sheet2, False, file, name_institute)
        if (' Área(s) Responsable de su Ejecución' == cell.value or 'ÁREA(S) RESPONSABLE(S) DE LA EJECUCIÓN DEL CONTRATO' == cell.value or
            'ÁREA(S) RESPONSABLE(S) QUE GENERA(N)  POSEE(N)  PUBLICA(N) Y ACTUALIZAN LA INFORMACIÓN' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 31, sheet2, False, file, name_institute)
        if (' Objeto Del Contrato' == cell.value or 'OBJETO DEL CONTRATO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 32, sheet2, False, file, name_institute)
        if (' Fecha de Inicio Del Plazo de Entrega O Ejecución' == cell.value or 'FECHA DE INICIO DEL PLAZO DE ENTREGA O EJECUCIÓN DE SERVICIOS CONTRATADOS U OBRA PÚBLICA' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 33, sheet2, False, file, name_institute)
        if (' Fecha de Término Del Plazo de Entrega O Ejecución' == cell.value or 'FECHA DE TÉRMINO DEL PLAZO DE ENTREGA O EJECUCIÓN DE SERVICIOS U OBRA PÚBLICA' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 34, sheet2, False, file, name_institute)
        if (' Hipervínculo Al Documento Del Contrato Y Anexos, en Versión Pública, en su Caso' == cell.value or
            'HIPERVÍNCULO AL DOCUMENTO DEL CONTRATO Y ANEXOS  VERSIÓN PÚBLICA SI ASÍ CORRESPONDE' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 35, sheet2, False, file, name_institute)
        if ('ESTATUS CONTRATO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 36, sheet2, False, file, name_institute)
        if (' Tipo de Fondo de Participación O Aportación Respectiva' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 37, sheet2, False, file, name_institute)
        if (' Breve Descripción de La Obra Pública, en su Caso' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 38, sheet2, False, file, name_institute)
        if (' Etapa de La Obra Pública Y/o Servicio de La Misma (catálogo)' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 39, sheet2, False, file, name_institute)
        if (' Mecanismos de Vigilancia Y Supervisión de La Ejecución, en su Caso' == cell.value or
            'MECANISMOS DE VIGILANCIA Y SUPERVISIÓN CONTRATOS' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 40, sheet2, False, file, name_institute)
        if (' Hipervínculo a Los Informes de Avance Financiero, en su Caso' == cell.value or
            'HIPERVÍNCULO A LOS INFORMES DE AVANCE FINANCIERO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 41, sheet2, False, file, name_institute)
        if (' Monto Total de Garantías Y/o Contragarantías, en Caso de Que Se Otorgaran durante El Procedimiento' == cell.value or
            'MONTO TOTAL DE GARANTÍAS Y/O CONTRAGARANTÍAS  EN CASO DE QUE SE OTORGARAN DURANTE EL PROCEDIMIENTO' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 42, sheet2, False, file, name_institute)
        if ('Institucion' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 43, sheet2, False, file, name_institute)
        if ('Municipio/Estado' == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 44, sheet2, False, file, name_institute)
        if (" Carácter Del Procedimiento (catálogo)" == cell.value):
            insertar(row_count, cur_col, maxima_row, sheet, row, 45, sheet2, False, file, name_institute)
        insertar(row_count, cur_col, maxima_row, sheet, row, 46, sheet2, False, file, name_institute)
    transparencia_excel.save("transparencia.xlsx")
def parse_home():
    for file in os.listdir():
        if file in ['.DS_Store']:
            continue 
        if file in ['transparencia.xlsx']:
            continue
        print("{}/{}".format(path,file))
        try:
            book = open_workbook("{}/{}".format(path,file),on_demand=True)
        except:
            print('este no se pudo abrir {}'.format(file))
            continue
        # data = open_file.values
        sheet = book.sheet_by_index(0)

        row_count = sheet.nrows
        print(row_count)
        col_count = sheet.ncols
        # maxima_columna = sheet_transparencia_excel.max_column
        maxima_row = sheet_transparencia_excel.max_row
        transparencia_excel.save("transparencia.xlsx")
        name = sheet.cell(0, 1)
        row = fila_iniciar(sheet, row_count)
        iter_columnas(col_count , sheet, maxima_row, row_count, row, book, False, file, name.value)
def run():
    parse_home()
if __name__ == '__main__':
    run()