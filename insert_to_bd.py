from bd import conexion
import csv, ipdb, psycopg2, os
from openpyxl import load_workbook
from datetime import datetime
import re

path = "/app/excels"
os.chdir(path)
transparencia_excel = load_workbook("{}/{}".format(path,'transparencia.xlsx'))
sheet_transparencia_excel = transparencia_excel["Sheet"]

def consultarId(tabla, propiedad, valor):
    consulta_id = ''
    try:
        with conexion.cursor() as cursor:
            if valor != None:
                consulta_id = "select id from {} where {}='{}';".format(tabla,propiedad,valor)
            else:
                consulta_id = "select id from {} where {}='{}';".format(tabla,propiedad,'')
            cursor.execute(consulta_id)
            # Con fetchall traemos todas las filas
            retorno = cursor.fetchall()
            # print(retorno)
            if len(retorno)>0:
                return retorno[0][0]
            else:
                return None
    except psycopg2.Error as e:
        print(consulta_id)
        print("Ocurrió un error al consultar con where: ", e)
        conexion.close()
    else:
        conexion.close()

def insertar(tabla, insert, cerrar_conexion):
    try:
        with conexion.cursor() as cursor:
            cursor.execute(insert)
        conexion.commit()  # Si no haces commit, los cambios no se guardan
    except psycopg2.Error as e:
        print(insert)
        print("Ocurrió un error al insertar: ", e)
        conexion.close()
    finally:
        if cerrar_conexion:
            print('cerro conexion')
            conexion.close()

def run():
    print(sheet_transparencia_excel.max_row)
    lista =[]
    duplicados = []
    # insert = "insert into transparencia (code_id,full_name_id,ejercicio_id,tipo_procedimiento_id,tipo_materia_id,folio,area_solicitud_id,area_contratante_id,numero_identifique_contrato,monto_sin_impuesto,monto_con_impuesto,tipo_moneda_id,tipo_cambio_id,forma_pago_id,origen_recurso_id,fuente_financiamiento_id,lugar,descripcion_obra,fecha_convocatoria,fecha_contrato,hiper_comunicado,hiper_dictamenes,hiper_convocatoria,hiper_fallo_junta,hiper_present_propues,descripcion_justifican,area_responsable_id,objeto_contrato,fecha_entrega_ejecucion,fecha_termino_ejecucion,hiper_contrato_publico,tipo_fondo_id,breve_descripcion_obra_publica,etapa_obra_id,mecanismo_vigilancia_id,hiper_informes_avance_financiero,monto_total_garantia,institute_id,estado_municipio,caracterde_procedimiento_id,name_file,usuario_registro_id, created, modified ) VALUES"
    codigo=full_name_id=ejercicio_id=tipo_procedimiento_id=materia_tipo_id=folio=area_solicitante_id=area_contratante_id=numero_identifique_contrato=monto_sin=monto_con=tipo_moneda_id=tipo_cambio_id=forma_pago_id=origen_recurso_id=fuente_financiamiento_id=lugar_obra=fecha_convocatoria=fecha_contrato=hiperviculo_comunicado=hiperviculo_dictamen=hiperviculo_convocatoria=hiperviculo_fallo=hiperviculo_presentacion=area_responsable_id=fecha_inicio_plazo=fecha_termino_plazo=hiperviculo_contrato=tipo_fondo_id=etapa_obra_id=mecanismo_vigilancia_id=hiperviculo_avance=monto_garantia=estado=caracter_procedimiento_id=file_name=institute_id = ''
    descripcion_justificaci = descripcion_obra = objeto_contrato = breve_descripcion_obra = ""
    nombres=last_name=surname=denominacion=rfc=""
    for row in range(2, (sheet_transparencia_excel.max_row + 1)):
        validar_codigo = sheet_transparencia_excel.cell(row= row, column = 1)
        duplicado = consultarId('transparencia', 'code_id',validar_codigo.value)
        if duplicado:
            duplicados.append(validar_codigo.value)
            print(duplicados)
            continue
        salto_de_linea = False
        if (isinstance(sheet_transparencia_excel.cell(row= row, column = 6).value,float) or isinstance(sheet_transparencia_excel.cell(row= row, column = 6).value,int)):
            salto_de_linea = True
        if (isinstance(sheet_transparencia_excel.cell(row= row, column = 7).value,float) or isinstance(sheet_transparencia_excel.cell(row= row, column = 7).value,int)):
            salto_de_linea = True
        if (isinstance(sheet_transparencia_excel.cell(row= row, column = 8).value,float) or isinstance(sheet_transparencia_excel.cell(row= row, column = 8).value,int)):
            salto_de_linea = True
        if (isinstance(sheet_transparencia_excel.cell(row= row, column = 9).value,float) or isinstance(sheet_transparencia_excel.cell(row= row, column = 9).value,int)):
            salto_de_linea = True
        if (isinstance(sheet_transparencia_excel.cell(row= row, column = 10).value,float) or isinstance(sheet_transparencia_excel.cell(row= row, column = 10).value,int)):
            salto_de_linea = True
        if (sheet_transparencia_excel.cell(row= row, column = 6).value == None and
            sheet_transparencia_excel.cell(row= row, column = 7).value == None and
            sheet_transparencia_excel.cell(row= row, column = 8).value == None and
            sheet_transparencia_excel.cell(row= row, column = 9).value == None and
            sheet_transparencia_excel.cell(row= row, column = 10).value == None ):
            salto_de_linea = True
        if salto_de_linea:
            continue
        
        instituto = sheet_transparencia_excel.cell(row= row, column = 45).value
        if not instituto in lista: # Imprime lo de abajo
            print(instituto)
            lista.append(instituto)
        for col in range(1, 11):
            cell = sheet_transparencia_excel.cell(row= row, column = col)
            if col == 1:
                # codigos.append(cell.value)
                codigo = cell.value
                # full_name_id = consultarId('full_name', 'code_id',cell.value)
                # if full_name_id == None:
                #     full_name_id = consultarId('full_name', 'code_id',0)
                # full_name.append(full_name_id)

            # if col == 2:
            #     if cell.value == None:
            #         ejercicio_id = consultarId('ejercicio', 'name',1970)
            #     else:
            #         ejercicio_id = consultarId('ejercicio', 'name',cell.value)
            #     # ejercicios.append(ejercicio_id)
            # if col == 3:
            #     # if not cell.value in lista: # Imprime lo de abajo
            #     #     lista.append(cell.value)
            #     tipo_procedimiento_id = consultarId('tipo_procedimiento', 'name', cell.value)
            #     if tipo_procedimiento_id == None:
            #         tipo_procedimiento_id = consultarId('tipo_procedimiento', 'name','')
            #     # tipo_procedimientos.append(tipo_procedimiento_id)
            # if col ==4:
            #     materia_tipo_id = consultarId('materia_tipo', 'name',cell.value)
            #     if materia_tipo_id == None:
            #         materia_tipo_id = consultarId('materia_tipo', 'name','')
            #     # materias.append(materia_tipo_id)
            # if col ==5:
            #     folio = cell.value
            #     # folios.append(cell.value)
            if col == 6:
                """Nombres """
                if cell.value == None:
                    nombres= ''
                else:
                    nombres = cell.value.strip()
                    
            if col ==7:
                """Last name """
                if cell.value == None:
                    last_name= ''
                else:
                    last_name = cell.value.strip()

            if col ==8:
                """Surname """
                if cell.value == None:
                    surname= ''
                else:
                    surname = cell.value.strip()

            if col ==9:
                """Razon social """
                if cell.value == None:
                    denominacion= ''
                else:
                    denominacion = cell.value.strip().replace("'","''",1)

            if col ==10:
                """RFC """
                if cell.value == None:
                    rfc= ''
                else:
                    rfc = cell.value.strip()
                    
            # if col ==11:
            #     area_solicitante_id = consultarId('area_solicitante', 'name',cell.value)
            #     if area_solicitante_id == None:
            #         area_solicitante_id = consultarId('area_solicitante', 'name','')
            #     # areas_solicitantes.append(area_solicitante_id)
            # if col ==12:
            #     area_contratante_id = consultarId('area_contratante', 'name',cell.value)
            #     if area_contratante_id == None:
            #         area_contratante_id = consultarId('area_contratante', 'name','')
            #     # areas_contratantes.append(area_contratante_id)
            # if col ==13:
            #     if cell.value == None:
            #         # numeros_identifique_contrato.append('')
            #         numero_identifique_contrato = ''
            #     else:
            #         # numeros_identifique_contrato.append(cell.value)
            #         numero_identifique_contrato = cell.value
            # if col ==14:
            #     if (isinstance(cell.value,float) or isinstance(cell.value,int)):
            #         # montos_sin.append(cell.value)
            #         monto_sin = cell.value
            #     else:
            #         # montos_sin.append(0)
            #         monto_sin = 0
            # if col ==15:
            #     if (isinstance(cell.value,float) or isinstance(cell.value,int)):
            #         # montos_con.append(cell.value)
            #         monto_con = cell.value
            #     else:
            #         # montos_con.append(0)
            #         monto_con = 0
            # if col ==16:
            #     tipo_moneda_id = consultarId('tipo_moneda', 'name',cell.value)
            #     if tipo_moneda_id == None:
            #         tipo_moneda_id = consultarId('tipo_moneda', 'name','')
            #     # tipos_monedas.append(tipo_moneda_id)
            # if col ==17:
            #     tipo_cambio_id = consultarId('tipo_cambio', 'name',cell.value)
            #     if tipo_cambio_id == None:
            #         tipo_cambio_id = consultarId('tipo_cambio', 'name','')
            #     # tipos_cambios.append(tipo_cambio_id)
            # if col ==18:
            #     forma_pago_id = consultarId('forma_pago', 'name',cell.value)
            #     if forma_pago_id == None:
            #         forma_pago_id = consultarId('forma_pago', 'name','')
            #     # formas_pago.append(forma_pago_id)
            # if col ==19:
            #     origen_recurso_id = consultarId('origen_recurso', 'name',cell.value)
            #     if origen_recurso_id == None:
            #         origen_recurso_id = consultarId('origen_recurso', 'name','')
            #     # origen_recursos.append(origen_recurso_id)
            # if col ==20:
            #     fuente_financiamiento_id = consultarId('fuente_financiamiento', 'name',cell.value)
            #     if fuente_financiamiento_id == None:
            #         fuente_financiamiento_id = consultarId('fuente_financiamiento', 'name','')
            #     # fuentes_financiamiento.append(fuente_financiamiento_id)
            # if col ==21:
            #     if cell.value == None:
            #         # lugar_obras.append('')
            #         lugar_obra = ''
            #     else:
            #         # lugar_obras.append(cell.value)
            #         lugar_obra = cell.value
            # if col ==22:
            #     if cell.value == None:
            #         # descripciones_obras.append('')
            #         descripcion_obra = ''
            #     else:
            #         # descripciones_obras.append(cell.value)
            #         descripcion_obra = re.sub("\'","",cell.value)

            # if col ==23:
            #     if cell.value == None:
            #         # fechas_convocatorias.append('1970-01-01')
            #         fecha_convocatoria = '1970-01-01'
            #     else:
            #         fecha_convocatoria = datetime.strptime(cell.value, '%d/%m/%Y').strftime('%Y-%m-%d')
            #         # fechas_convocatorias.append(fecha_convocatoria)
            # if col ==24:
            #     if cell.value == None:
            #         # fechas_contratos.append('1970-01-01')
            #         fecha_contrato = '1970-01-01'
            #     else:
            #         fecha_contrato =datetime.strptime(cell.value, '%d/%m/%Y').strftime('%Y-%m-%d')
            #         # fechas_contratos.append(fecha_contrato)
            # if col ==25:
            #     if cell.value == None:
            #         # hiperviculos_comunicado.append('')
            #         hiperviculo_comunicado = ''
            #     else:
            #         # hiperviculos_comunicado.append(cell.value)
            #         hiperviculo_comunicado = cell.value
            # if col ==26:
            #     if cell.value == None:
            #         # hiperviculos_dictamen.append('')
            #         hiperviculo_dictamen = ''
            #     else:
            #         # hiperviculos_dictamen.append(cell.value)
            #         hiperviculo_dictamen = cell.value
            # if col ==27:
            #     if cell.value == None:
            #         # hiperviculos_convocatoria.append('')
            #         hiperviculo_convocatoria = ''
            #     else:
            #         # hiperviculos_convocatoria.append(cell.value)
            #         hiperviculo_convocatoria = cell.value
            # if col ==28:
            #     if cell.value == None:
            #         # hiperviculos_fallo.append('')
            #         hiperviculo_fallo = ''
            #     else:
            #         # hiperviculos_fallo.append(cell.value)
            #         hiperviculo_fallo = cell.value
            # if col ==29:
            #     if cell.value == None:
            #         # hiperviculos_presentacion.append('')
            #         hiperviculo_presentacion = ''
            #     else:
            #         # hiperviculos_presentacion.append(cell.value)
            #         hiperviculo_presentacion = cell.value
            # if col ==30:
            #     if cell.value == None:
            #         # descripcion_justificacion.append('')
            #         descripcion_justificaci = ''
            #     else:
            #         # descripcion_justificacion.append(cell.value)
            #         descripcion_justificaci = re.sub("\'","",cell.value)
            # if col ==31:
            #     area_responsable_id = consultarId('area_responsable', 'name',cell.value)
            #     if area_responsable_id == None:
            #         area_responsable_id = consultarId('area_responsable', 'name','')
            #     # areas_responsables.append(area_responsable_id)
            # if col ==32:
            #     if cell.value == None:
            #         # objetos_contratos.append('')
            #         objeto_contrato = ''
            #     else:
            #         # objetos_contratos.append(cell.value)
            #         objeto_contrato = re.sub("\'","",cell.value)
            # if col ==33:
            #     if cell.value == None:
            #         # fechas_inicio_plazo.append('1970-01-01')
            #         fecha_inicio_plazo = '1970-01-01'
            #     else:
            #         fecha_inicio_plazo =datetime.strptime(cell.value, '%d/%m/%Y').strftime('%Y-%m-%d')
            #         # fechas_inicio_plazo.append(fecha_inicio_plazo)
            # if col ==34:
            #     if cell.value == None:
            #         # fechas_termino_plazo.append('1970-01-01')
            #         fecha_termino_plazo = '1970-01-01'
            #     else:
            #         fecha_termino_plazo =datetime.strptime(cell.value, '%d/%m/%Y').strftime('%Y-%m-%d')
            #         # fechas_termino_plazo.append(fecha_termino_plazo)
            # if col ==35:
            #     if cell.value == None:
            #         # hiperviculos_contrato.append('')
            #         hiperviculo_contrato = ''
            #     else:
            #         # hiperviculos_contrato.append(cell.value)
            #         hiperviculo_contrato = cell.value
            # if col ==36:
            #     tipo_fondo_id = consultarId('tipo_fondo', 'name',cell.value)
            #     if tipo_fondo_id == None:
            #         tipo_fondo_id = consultarId('tipo_fondo', 'name','')
            #     # tipos_fondos.append(tipo_fondo_id)
            # if col ==37:
            #     if cell.value == None or cell.value == 0:
            #         # breves_descripcion_obra.append('')
            #         breve_descripcion_obra = ''
            #     else:
            #         # breves_descripcion_obra.append(cell.value)
            #         breve_descripcion_obra = re.sub("\'","",cell.value)
            # if col ==38:
            #     etapa_obra_id = consultarId('etapa_obra', 'name',cell.value)
            #     if etapa_obra_id == None:
            #         etapa_obra_id = consultarId('etapa_obra', 'name','')
            #     # etapas_obra.append(etapa_obra_id)
            # if col ==39:
            #     mecanismo_vigilancia_id = consultarId('mecanismo_vigilancia', 'name',cell.value)
            #     if mecanismo_vigilancia_id == None:
            #         mecanismo_vigilancia_id = consultarId('mecanismo_vigilancia', 'name','')
            #     # mecanismos_vigilancia.append(mecanismo_vigilancia_id)
            # if col ==40:
            #     if cell.value == None:
            #         # hiperviculos_avance_fianciero.append('')
            #         hiperviculo_avance = ''
            #     else:
            #         # hiperviculos_avance_fianciero.append(cell.value)
            #         hiperviculo_avance = cell.value
            # if col ==41:
            #     if (isinstance(cell.value,float) or isinstance(cell.value,int)):
            #         # montos_garantias.append(cell.value)
            #         monto_garantia = cell.value
            #     else:
            #         # montos_garantias.append(0)
            #         monto_garantia = 0
            # if col ==42:
            #     if cell.value == None:
            #         # estados.append('')
            #         estado = ''
            #     else:
            #         # estados.append(cell.value)
            #         estado = cell.value
            # if col ==43:
            #     caracter_procedimiento_id = consultarId('caracter_procedimiento', 'name',cell.value)
            #     if caracter_procedimiento_id == None:
            #         caracter_procedimiento_id = consultarId('mecanismo_vigilancia', 'name','')
            #     # caracter_procedimientos.append(caracter_procedimiento_id)
            # if col ==44:
            #     # files.append(cell.value)
            #     file_name= cell.value
            #     if not cell.value in lista: # Imprime lo de abajo
            #         print(cell.value)
            #         lista.append(cell.value)
            # if col ==45:
            #     # institute_id = consultarId('institute', 'name',cell.value.strip())
            #     # if institute_id == None:
            #     #     institute_id = consultarId('institute', 'name','')
            #     # institutos.append(institute_id)
            #     if not cell.value in lista: # Imprime lo de abajo
            #         print(cell.value)
            #         lista.append(cell.value)
        # insert = "insert into transparencia (code_id,full_name_id,ejercicio_id,tipo_procedimiento_id,tipo_materia_id,folio,area_solicitud_id,area_contratante_id,numero_identifique_contrato,monto_sin_impuesto,monto_con_impuesto,tipo_moneda_id,tipo_cambio_id,forma_pago_id,origen_recurso_id,fuente_financiamiento_id,lugar,descripcion_obra,fecha_convocatoria,fecha_contrato,hiper_comunicado,hiper_dictamenes,hiper_convocatoria,hiper_fallo_junta,hiper_present_propues,descripcion_justifican,area_responsable_id,objeto_contrato,fecha_entrega_ejecucion,fecha_termino_ejecucion,hiper_contrato_publico,tipo_fondo_id,breve_descripcion_obra_publica,etapa_obra_id,mecanismo_vigilancia_id,hiper_informes_avance_financiero,monto_total_garantia,institute_id,estado_municipio,caracterde_procedimiento_id,name_file,usuario_registro_id, created, modified ) VALUES ({},{},{},{},{},'{}',{},{},'{}',{},{},{},{},{},{},{},'{}','{}','{}','{}','{}','{}','{}','{}','{}','{}',{},'{}','{}','{}','{}',{},'{}',{},{},'{}',{},{},'{}',{},'{}',1, current_timestamp, current_timestamp);".format(
        #     codigo,
        #     full_name_id,
        #     ejercicio_id,
        #     tipo_procedimiento_id,
        #     materia_tipo_id,
        #     folio,
        #     area_solicitante_id,
        #     area_contratante_id,
        #     numero_identifique_contrato,
        #     monto_sin,
        #     monto_con,
        #     tipo_moneda_id,
        #     tipo_cambio_id,
        #     forma_pago_id,
        #     origen_recurso_id,
        #     fuente_financiamiento_id,
        #     lugar_obra,
        #     descripcion_obra,
        #     fecha_convocatoria,
        #     fecha_contrato,
        #     hiperviculo_comunicado,
        #     hiperviculo_dictamen,
        #     hiperviculo_convocatoria,
        #     hiperviculo_fallo,
        #     hiperviculo_presentacion,
        #     descripcion_justificaci,
        #     area_responsable_id,
        #     objeto_contrato,
        #     fecha_inicio_plazo,
        #     fecha_termino_plazo,
        #     hiperviculo_contrato,
        #     tipo_fondo_id,
        #     breve_descripcion_obra,
        #     etapa_obra_id,
        #     mecanismo_vigilancia_id,
        #     hiperviculo_avance,
        #     monto_garantia,
        #     institute_id,
        #     estado,
        #     caracter_procedimiento_id,
        #     file_name)
        insert = "insert into full_name (names, last_name, surname, razon_social, rfc, code_id, usuario_registro_id, created, modified) VALUES('{}','{}','{}','{}','{}',{},1, current_timestamp, current_timestamp)".format(
            nombres,
            last_name,
            surname,
            denominacion,
            rfc,
            codigo)
        if(row != sheet_transparencia_excel.max_row):
            insertar('transparencia', insert, False)
        if(row == sheet_transparencia_excel.max_row):
            print(insert)
            insertar('transparencia', insert, True)

if __name__ == '__main__':
    run()